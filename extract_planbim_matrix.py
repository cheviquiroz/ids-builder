"""
Extrae de la Matriz PlanBIM V3.0 los parámetros reales de las entidades de
Estructura (Columnas, Vigas, Losas, Fundaciones, Muros) para las etapas
Diseño Conceptual (DC) y Diseño Básico (DB).

Requiere: pip install openpyxl

Uso: python3 extract_planbim_matrix.py
Genera: planbim_matrix_extracted.json
"""

import json
import sys

import openpyxl

SOURCE_FILE = "Matriz-de-Informacion-de-Entidades_Version_3.0.xlsx"
SHEET_NAME = "2. Lista parámetros por Entidad"
OUTPUT_FILE = "planbim_matrix_extracted.json"

# Entidades IFC de Estructura que nos interesan.
ENTIDADES_ESTRUCTURA = {"IfcColumn", "IfcBeam", "IfcSlab", "IfcFooting", "IfcWall"}

# Etapas de interés (EAIM = Etapa de Aplicación de la Información Mínima).
ETAPAS_INTERES = {"DC", "DB"}


def es_valor_valido(value):
    """Un valor de celda cuenta si no está vacío y no es literalmente 'No aplica'."""
    if value is None:
        return False
    text = str(value).strip()
    return text != "" and text.lower() != "no aplica"


def clasificar_categoria(pset, propiedad_ifc, nombre_param):
    """
    Clasifica cada parámetro en una categoría humana, a partir de patrones
    reales observados en la Matriz (nombre del Pset y de la propiedad IFC).
    """
    pset_val = pset if es_valor_valido(pset) else ""
    prop_val = propiedad_ifc if es_valor_valido(propiedad_ifc) else ""
    nombre = (nombre_param or "").strip().lower()

    if pset_val.startswith("Qto_"):
        return "dimensiones"
    if prop_val in ("Slope", "PitchAngle"):
        return "dimensiones"
    if prop_val == "FireRating":
        return "fuego"
    if prop_val == "LoadBearing":
        return "cargas"
    if prop_val in ("YieldStrength", "ElasticModulus", "CompressiveStrength", "TensileStrength"):
        return "resistencia"
    if prop_val == "ThermalTransmittance" or "valor u" in nombre or "térmic" in nombre:
        return "termico"
    if nombre == "material" or "material" in prop_val.lower():
        return "materiales"
    return "clasificacion"


def nombre_tecnico(atributo_ifc, propiedad_ifc, nombre_param):
    """
    El identificador técnico real de un parámetro: prioriza la propiedad de
    Pset (columna 'Propiedad IFC', p.ej. 'Height', 'LoadBearing') porque es
    la más específica; si no aplica, usa el atributo nativo IFC (p.ej.
    'Name', 'PredefinedType'); si tampoco aplica, cae al nombre en español.
    """
    if es_valor_valido(propiedad_ifc):
        return propiedad_ifc.strip()
    if es_valor_valido(atributo_ifc):
        return atributo_ifc.strip()
    return nombre_param.strip()


def main():
    try:
        wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"❌ Error: no se encontró '{SOURCE_FILE}' en la carpeta actual.")
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Error: la hoja '{SHEET_NAME}' no existe. Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    print("✅ Excel abierto")

    # Estructura: {ifc: {etapa: {ndi: [params]}}}
    matriz_filtrada = {}
    # Evita filas duplicadas (misma entidad/etapa/ndi/parámetro) presentes en el Excel original.
    vistos = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        entidad_ifc = row[1]
        eaim = row[2]
        tdi = row[3]
        ndi = row[4]
        nombre_param = row[5]
        atributo_ifc = row[6]
        pset = row[8]
        propiedad_ifc = row[9]

        if entidad_ifc not in ENTIDADES_ESTRUCTURA:
            continue
        if eaim not in ETAPAS_INTERES:
            continue
        if not nombre_param or str(nombre_param).strip() == "":
            continue

        english = nombre_tecnico(atributo_ifc, propiedad_ifc, nombre_param)
        spanish = nombre_param.strip()

        dedupe_key = (entidad_ifc, eaim, ndi, spanish, english)
        if dedupe_key in vistos:
            continue
        vistos.add(dedupe_key)

        matriz_filtrada.setdefault(entidad_ifc, {}).setdefault(eaim, {}).setdefault(ndi, [])

        matriz_filtrada[entidad_ifc][eaim][ndi].append(
            {
                "spanish": spanish,
                "english": english,
                "tdi": tdi,
                "ndi": ndi,
                "pset": pset.strip() if es_valor_valido(pset) else None,
                "category": clasificar_categoria(pset, propiedad_ifc, nombre_param),
            }
        )

    print("✅ Matriz filtrada\n")

    for ifc in sorted(matriz_filtrada.keys()):
        print(f"{ifc}:")
        for etapa in sorted(matriz_filtrada[ifc].keys()):
            ndi_dict = matriz_filtrada[ifc][etapa]
            for ndi in sorted(ndi_dict.keys()):
                print(f"  {etapa} → {ndi}: {len(ndi_dict[ndi])} parámetros")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(matriz_filtrada, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Guardado: {OUTPUT_FILE}")
    print(f"✅ Total entidades: {len(matriz_filtrada)}")


if __name__ == "__main__":
    main()
